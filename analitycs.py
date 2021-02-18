import pandas as pd
import math
from openpyxl.formatting.rule import DataBarRule, Rule
from openpyxl.styles import PatternFill, Font, numbers, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
import statistics as stats
import matplotlib
import matplotlib.cm
import matplotlib.colors
import matplotlib.pyplot as plt
import numpy as np
import squarify
import statsmodels.formula.api as smf
from datetime import datetime
from openpyxl import load_workbook, drawing
from os import listdir, remove

csv_collection = 'collection.csv'
excel_file = 'collection.xlsx'


def highlight_excel():
    workbook = load_workbook(filename=excel_file)
    sheet = workbook.active
    row_count = sheet.max_row
    dimensions = sheet.dimensions
    sheet.auto_filter.ref = dimensions

    red_background = PatternFill(bgColor='ff4d4d')
    yellow_background = PatternFill(bgColor='ffff80')
    green_background = PatternFill(bgColor='ccffcc')

    diff_style1 = DifferentialStyle(fill=red_background)
    diff_style2 = DifferentialStyle(fill=yellow_background)
    diff_style3 = DifferentialStyle(fill=green_background)

    rule3 = Rule(type="expression", dxf=diff_style3)
    rule3.formula = ["$D1>=7"]
    sheet.conditional_formatting.add('A1:C' + str(row_count), rule3)
    sheet.conditional_formatting.add('E1:J' + str(row_count), rule3)

    rule1 = Rule(type="expression", dxf=diff_style1)
    rule1.formula = ["$D1<5"]
    sheet.conditional_formatting.add('A1:C' + str(row_count), rule1)
    sheet.conditional_formatting.add('E1:J' + str(row_count), rule1)

    rule2 = Rule(type="expression", dxf=diff_style2)
    rule2.formula = (["$D1<7"])
    sheet.conditional_formatting.add('A1:C' + str(row_count), rule2)
    sheet.conditional_formatting.add('E1:J' + str(row_count), rule2)

    data = pd.read_csv(csv_collection)
    start_value = min(data['SCORE'])
    end_value = max(data['SCORE'])

    data_bar_rule = DataBarRule(start_type="num", start_value=start_value,
                                end_type="num", end_value=end_value, color='53ff4d')
    sheet.conditional_formatting.add('D2:D' + str(row_count), data_bar_rule)

    red_text = Font(color='a70000')
    green_text = Font(color='00bb00')

    for element in range(2, row_count + 1):
        sheet['F' + str(element)].number_format = numbers.FORMAT_CURRENCY_USD
        sheet['G' + str(element)].number_format = numbers.FORMAT_CURRENCY_USD
        sheet['H' + str(element)].number_format = numbers.FORMAT_CURRENCY_USD
        cell_value = sheet.cell(element, 8).value
        if int(cell_value) < 0:
            sheet.cell(element, 8).font = red_text
        else:
            sheet.cell(element, 8).font = green_text

    workbook.save(excel_file)


def convert_csv_into_excel():
    if excel_file in listdir('.'):
        remove(excel_file)
    data = pd.read_csv(csv_collection)
    data.to_excel(excel_file, sheet_name='Info', index=None, header=True, freeze_panes=(1, 0))


def border_cell(sheet, range_cell):
    thin = Side(border_style="thin", color="000000")  # Border style, color
    border = Border(left=thin, right=thin, top=thin, bottom=thin)  # Position of border
    for row in sheet[range_cell]:
        for cell in row:
            cell.border = border


def merge_cells(sheet, range_cell):
    sheet.merge_cells(range_cell)
    cell = sheet.cell(row=1, column=1)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def insert_rating_number_year(list1, sheet, column1, column2, column3, position, list2, list3):
    for element in range(0, len(list1)):
        sheet[column1 + str(position + element)] = list1[element]
        sheet[column2 + str(position + element)] = list2[element]
        sheet[column3 + str(position + element)] = round(list3[element], 3)


def save_and_insert_image(excel_sheet, name, cell):
    plt.savefig(name + '.png')
    img = drawing.image.Image(name + '.png')
    img.anchor = cell
    excel_sheet.add_image(img)


def charts():
    start = datetime.now()
    data = pd.read_csv(csv_collection)
    convert_csv_into_excel()
    workbook = load_workbook(filename=excel_file)
    sheets = [workbook.sheetnames]
    if 'Charts' not in sheets:
        chart_sheet = workbook.create_sheet("Charts&Analytics")
    else:
        chart_sheet = workbook.get_sheet_by_name('Charts&Analytics')
        workbook.remove_sheet(chart_sheet)
        chart_sheet = workbook.create_sheet("Charts&Analytics")

    # 1) Donut chart of percentage of films of each category
    chart_sheet['B1'] = 'MOVIES BY CATEGORY'
    plt.figure()
    data_grouped_by_category = data.groupby('CATEGORY').count()
    number_of_movies_by_category = [value for value in data_grouped_by_category['ID']]
    plt.pie(number_of_movies_by_category, labels=['Pendientes', 'Vistas'], autopct='%0.2f%%', startangle=90)
    circle = plt.Circle((0, 0), 0.75, fc='white')
    donut = plt.gcf()
    donut.gca().add_artist(circle)
    plt.axis('equal')
    plt.tight_layout()
    save_and_insert_image(chart_sheet, 'Donut chart of percentage of films of each category', 'A2')

    chart_sheet['J3'] = 'STORED MOVIES'
    chart_sheet['K3'] = len(data)
    chart_sheet['J4'] = 'VIEWED MOVIES'
    chart_sheet['K4'] = number_of_movies_by_category[1]
    chart_sheet['J5'] = 'PENDING MOVIES'
    chart_sheet['K5'] = number_of_movies_by_category[0]

    # 2) Radar chart of films of each category (released movies)
    plt.figure()
    released_movies = data
    for element in range(0, len(released_movies['SCORE'])):
        if released_movies['SCORE'][element] == 0:
            released_movies = released_movies.drop([element], axis=0)

    released_movies_grouped_by_category = released_movies.groupby('CATEGORY').mean()
    runtime_mean_by_category = [value for value in released_movies_grouped_by_category['RUNTIME']]
    max_runtime_mean_by_category = max(runtime_mean_by_category)
    gross_mean_by_category = [value for value in released_movies_grouped_by_category['GROSS']]
    max_gross_mean_by_category = max(gross_mean_by_category)
    budget_mean_by_category = [value for value in released_movies_grouped_by_category['BUDGET']]
    max_budget_mean_by_category = max(budget_mean_by_category)
    score_mean_by_category = [value for value in released_movies_grouped_by_category['SCORE']]
    max_score_mean_by_category = max(score_mean_by_category)

    df = pd.DataFrame({
        'group': ['Pendientes', 'Vistas'],
        'Runtime': [runtime_mean_by_category[0] / max_runtime_mean_by_category,
                    runtime_mean_by_category[1] / max_runtime_mean_by_category],
        'Score': [score_mean_by_category[0] / max_score_mean_by_category,
                  score_mean_by_category[1] / max_score_mean_by_category],
        'Gross': [gross_mean_by_category[0] / max_gross_mean_by_category,
                  gross_mean_by_category[1] / max_gross_mean_by_category],
        'Budget': [budget_mean_by_category[0] / max_budget_mean_by_category,
                   budget_mean_by_category[1] / max_budget_mean_by_category],
    })

    categories = list(df)[1:]
    number_of_categories = len(categories)

    angles = [n / float(number_of_categories) * 2 * math.pi for n in range(number_of_categories)]
    angles += angles[:1]

    ax = plt.subplot(111, polar=True)
    ax.set_theta_offset(math.pi / 2)
    ax.set_theta_direction(-1)

    plt.xticks(angles[:-1], categories)

    ax.set_rlabel_position(0)
    plt.yticks([0.25, 0.5, 0.75], ["0.25", "0.5", "0.75"], color="grey", size=7)
    plt.ylim(0, 1)

    values = df.loc[0].drop('group').values.flatten().tolist()
    values += values[:1]
    ax.plot(angles, values, linewidth=1, linestyle='solid', label="Pendientes")
    ax.fill(angles, values, 'b', alpha=0.1)

    values = df.loc[1].drop('group').values.flatten().tolist()
    values += values[:1]
    ax.plot(angles, values, linewidth=1, linestyle='solid', label="Vistas")
    ax.fill(angles, values, 'r', alpha=0.1)
    plt.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1))

    save_and_insert_image(chart_sheet, 'Radar chart of released films of each category', 'A28')

    chart_sheet['K31'] = 'Pendientes'
    chart_sheet['L31'] = 'Vistas'
    chart_sheet['J32'] = 'Mean Score'
    chart_sheet['J33'] = 'Mean Budget'
    chart_sheet['J34'] = 'Mean Gross'
    chart_sheet['J35'] = 'Mean Runtime'

    chart_sheet['K32'] = score_mean_by_category[0]
    chart_sheet['K33'] = budget_mean_by_category[0]
    chart_sheet['K34'] = gross_mean_by_category[0]
    chart_sheet['K35'] = runtime_mean_by_category[0]

    chart_sheet['L32'] = score_mean_by_category[1]
    chart_sheet['L33'] = budget_mean_by_category[1]
    chart_sheet['L34'] = gross_mean_by_category[1]
    chart_sheet['L35'] = runtime_mean_by_category[1]

    # 3) number of films per decade (bar)
    dates_label = ['80s', '90s', '00s', '10s', '20s']
    plt.figure()
    dates = data['YEAR']
    decade_80s = [element for element in dates if 1980 <= element < 1990]
    decade_90s = [element for element in dates if 1990 <= element < 2000]
    decade_00s = [element for element in dates if 2000 <= element < 2010]
    decade_10s = [element for element in dates if 2010 <= element < 2020]
    decade_20s = [element for element in dates if 2020 <= element]
    values = [len(decade_80s), len(decade_90s), len(decade_00s), len(decade_10s), len(decade_20s)]
    plt.bar(dates_label, values, color='#70b5ee', width=0.5)
    plt.xlabel("Decades")
    plt.ylabel("Nº of movies")
    save_and_insert_image(chart_sheet, 'number of films per decade', 'A54')

    # 4) number of movies per decade classified as seen or pending (bar)
    chart_sheet['B53'] = 'MOVIES PER DECADE'
    plt.figure()
    viewed_dataframe = data
    for element in range(0, len(viewed_dataframe['CATEGORY'])):
        if viewed_dataframe['CATEGORY'][element] == 'Pendiente':
            viewed_dataframe = viewed_dataframe.drop([element], axis=0)

    years_seen = viewed_dataframe['YEAR']
    decade80s_vista = [element for element in years_seen if 1980 <= element < 1990]
    decade90s_vista = [element for element in years_seen if 1990 <= element < 2000]
    decade00s_vista = [element for element in years_seen if 2000 <= element < 2010]
    decade10s_vista = [element for element in years_seen if 2010 <= element < 2020]
    decade20s_vista = [element for element in years_seen if 2020 <= element]

    pending_dataframe = data
    for element in range(0, len(pending_dataframe['CATEGORY'])):
        if pending_dataframe['CATEGORY'][element] == 'Vista':
            pending_dataframe = pending_dataframe.drop([element], axis=0)

    years_pending = pending_dataframe['YEAR']
    decade80s_pending = [element for element in years_pending if 1980 <= element < 1990]
    decade90s_pending = [element for element in years_pending if 1990 <= element < 2000]
    decade00s_pending = [element for element in years_pending if 2000 <= element < 2010]
    decade10s_pending = [element for element in years_pending if 2010 <= element < 2020]
    decade20s_pending = [element for element in years_pending if 2020 <= element]

    bars_vistas = [len(decade80s_vista), len(decade90s_vista), len(decade00s_vista),
                   len(decade10s_vista), len(decade20s_vista)]
    bars_pending = [len(decade80s_pending), len(decade90s_pending), len(decade00s_pending),
                    len(decade10s_pending), len(decade20s_pending)]

    bar_width = 0.25
    decades1 = np.arange(len(bars_vistas))
    decades2 = [x + bar_width for x in decades1]
    plt.bar(decades1, bars_vistas, color='#534bff', width=bar_width, label='Vistas')
    plt.bar(decades2, bars_pending, color='#7dadff', width=bar_width, label='Pendientes')
    plt.xticks([r + bar_width / 2 for r in range(len(bars_vistas))], dates_label)
    plt.xlabel("Decades")
    plt.ylabel("Nº of movies")
    plt.legend()
    save_and_insert_image(chart_sheet, 'Number of movies per decade classified as seen or pending', 'J54')

    chart_sheet['S56'] = 'DECADE'
    chart_sheet['T56'] = 'CATEGORY'
    chart_sheet['U56'] = 'MOVIES'

    chart_sheet['S57'] = '80s'
    chart_sheet['T57'] = 'Viewed'
    chart_sheet['U57'] = len(decade80s_vista)
    chart_sheet['T58'] = 'Pending'
    chart_sheet['U58'] = len(decade80s_pending)
    chart_sheet['T59'] = 'Total'
    chart_sheet['U59'] = len(decade_80s)

    chart_sheet['S60'] = '90s'
    chart_sheet['T60'] = 'Viewed'
    chart_sheet['U60'] = len(decade90s_vista)
    chart_sheet['T61'] = 'Pending'
    chart_sheet['U61'] = len(decade90s_pending)
    chart_sheet['T62'] = 'Total'
    chart_sheet['U62'] = len(decade_90s)

    chart_sheet['S63'] = '00s'
    chart_sheet['T63'] = 'Viewed'
    chart_sheet['U63'] = len(decade00s_vista)
    chart_sheet['T64'] = 'Pending'
    chart_sheet['U64'] = len(decade00s_pending)
    chart_sheet['T65'] = 'Total'
    chart_sheet['U65'] = len(decade_00s)

    chart_sheet['S66'] = '10s'
    chart_sheet['T66'] = 'Viewed'
    chart_sheet['U66'] = len(decade10s_vista)
    chart_sheet['T67'] = 'Pending'
    chart_sheet['U67'] = len(decade10s_pending)
    chart_sheet['T68'] = 'Total'
    chart_sheet['U68'] = len(decade_10s)

    chart_sheet['S69'] = '20s'
    chart_sheet['T69'] = 'Viewed'
    chart_sheet['U69'] = len(decade20s_vista)
    chart_sheet['T70'] = 'Pending'
    chart_sheet['U70'] = len(decade20s_pending)
    chart_sheet['T71'] = 'Total'
    chart_sheet['U71'] = len(decade_20s)

    # 5) Percentage of films of each decade (pie)
    plt.figure()
    y = np.array([len(decade_80s), len(decade_90s), len(decade_00s), len(decade_10s), len(decade_20s)])
    my_explode = [0, 0, 0, 0.2, 0]
    colors_pie = ['#ff9999', '#66b3ff', '#FFE873', '#99ff99', '#ffcc99']
    plt.pie(y, autopct='%.2f%%', labels=dates_label, explode=my_explode, shadow=True, colors=colors_pie)
    save_and_insert_image(chart_sheet, 'Pie diagram of percentage of films of each decade', 'A80')

    # 6) Number of films by year
    chart_sheet['B105'] = 'MOVIES PER YEAR'
    plt.figure()
    data_grouped_by_year = data.groupby('YEAR').count()
    number_movies = [element for element in data_grouped_by_year['SCORE']]
    lowest_number_movies = min(number_movies)
    highest_number_movies = max(number_movies)
    nota_media = stats.mean(number_movies)
    years = []
    for element in data['YEAR']:
        if element not in years:
            years.append(int(element))
    sorted_years = sorted(years, reverse=False)
    plt.axhline(nota_media, color="red", linewidth=1.5, linestyle="dashed")
    plt.axis([min(sorted_years) - 2, max(sorted_years) + 2, lowest_number_movies - 1, highest_number_movies + 1])
    plt.plot(sorted_years, number_movies, color="Slateblue", alpha=0.6)
    plt.xlabel("Years")
    plt.ylabel("Number of movies stored")
    plt.fill_between(sorted_years, number_movies, color="skyblue", alpha=0.4)
    plt.fill_between(sorted_years, number_movies, color="skyblue", alpha=0.2)
    save_and_insert_image(chart_sheet, 'Number of films by year', 'A106')

    # 7) Average film score by year
    plt.figure()
    data_grouped_by_year = data.groupby('YEAR').mean()
    mean_score_year = [element for element in data_grouped_by_year['SCORE']]
    minimum_score = min(mean_score_year)
    nota_media = stats.mean(mean_score_year)
    years = []
    for element in data['YEAR']:
        if element not in years:
            years.append(int(element))
    sorted_years = sorted(years, reverse=False)
    plt.axhline(nota_media, color="red", linewidth=1.5, linestyle="dashed")
    plt.axis([min(sorted_years) - 2, max(sorted_years) + 2, minimum_score - 1, 10])
    plt.plot(sorted_years, mean_score_year, color="green", alpha=0.6)
    plt.fill_between(sorted_years, mean_score_year, color="green", alpha=0.4)
    plt.fill_between(sorted_years, mean_score_year, color="green", alpha=0.2)
    plt.xlabel("Years")
    plt.ylabel("Average score")
    save_and_insert_image(chart_sheet, 'average film score', 'J106')

    chart_sheet['S107'] = 'YEAR'
    chart_sheet['T107'] = 'MOVIES'
    chart_sheet['U107'] = 'SCORE'
    chart_sheet['V107'] = 'YEAR'
    chart_sheet['W107'] = 'MOVIES'
    chart_sheet['X107'] = 'SCORE'
    chart_sheet['Y107'] = 'YEAR'
    chart_sheet['Z107'] = 'MOVIES'
    chart_sheet['AA107'] = 'SCORE'

    years80 = [element for element in sorted_years if '198' in str(element)]
    years90 = [element for element in sorted_years if '199' in str(element)]
    years00 = [element for element in sorted_years if '200' in str(element)]
    years10 = [element for element in sorted_years if '201' in str(element)]
    years20 = [element for element in sorted_years if '202' in str(element)]

    insert_rating_number_year(years80, chart_sheet, 'S', 'T', 'U', 108, number_movies, mean_score_year)
    insert_rating_number_year(years90, chart_sheet, 'S', 'T', 'U', 118, number_movies[10:20], mean_score_year[10:20])
    insert_rating_number_year(years00, chart_sheet, 'V', 'W', 'X', 108, number_movies[20:30], mean_score_year[20:30])
    insert_rating_number_year(years10, chart_sheet, 'V', 'W', 'X', 118, number_movies[30:40], mean_score_year[30:40])
    insert_rating_number_year(years20, chart_sheet, 'Y', 'Z', 'AA', 108, number_movies[40:], mean_score_year[40:])

    # 8) Histogram of scores
    chart_sheet['B132'] = 'Histogram of scores'
    plt.figure()
    scores = data['SCORE']
    plt.hist(scores, bins=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10], ec='black')
    plt.xlabel('Score')
    plt.ylabel('Frequency')
    save_and_insert_image(chart_sheet, 'Histogram of scores', 'A133')

    chart_sheet['K141'] = 'Score'
    chart_sheet['L141'] = 'Movies'

    zeros = [element for element in data['SCORE'] if int(element) == 0]
    ones = [element for element in data['SCORE'] if int(element) == 1]
    twos = [element for element in data['SCORE'] if int(element) == 2]
    threes = [element for element in data['SCORE'] if int(element) == 3]
    fours = [element for element in data['SCORE'] if int(element) == 4]
    fives = [element for element in data['SCORE'] if int(element) == 5]
    sixes = [element for element in data['SCORE'] if int(element) == 6]
    sevens = [element for element in data['SCORE'] if int(element) == 7]
    eights = [element for element in data['SCORE'] if int(element) == 8]
    nines = [element for element in data['SCORE'] if int(element) == 9]
    tens = [element for element in data['SCORE'] if int(element) == 10]

    scores_movies = [len(zeros), len(ones), len(twos), len(threes), len(fours), len(fives), len(sixes), len(sevens),
                     len(eights), len(nines), len(tens)]
    scores_list = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

    for element in range(0, len(scores_list)):
        chart_sheet['K' + str(142 + element)] = scores_list[element]
        chart_sheet['L' + str(142 + element)] = scores_movies[element]

    # 9) 3d Budget-to-Gross Ratio
    cinema_csv_file = data
    for element in range(0, len(cinema_csv_file['BENEFIT'])):
        if cinema_csv_file['BUDGET'][element] == 0 or cinema_csv_file['GROSS'][element] == 0:
            cinema_csv_file = cinema_csv_file.drop([element], axis=0)
    budget = cinema_csv_file['BUDGET']
    gross = cinema_csv_file['GROSS']
    plt.figure()
    chart_sheet['B159'] = 'Budget-to-Gross Ratio'
    ax = plt.axes(projection='3d')
    ax.set_xlabel('Budget')
    ax.set_ylabel('Gross')
    ax.set_zlabel('Benefit')
    benefit = cinema_csv_file['BENEFIT']
    ax.scatter3D(budget, gross, benefit, c=benefit, cmap='winter')
    save_and_insert_image(chart_sheet, '3d Budget-to-Gross Ratio', 'A160')

    # 10) Budget-to-Gross Ratio by decade
    plt.figure()
    chart_sheet['T164'] = 'Decade'
    chart_sheet['U164'] = 'Movies'

    chart_sheet['T165'] = '80s'
    data_copy_80s = cinema_csv_file
    for element in range(0, len(data)):
        if element in data_copy_80s['YEAR'] and data_copy_80s['YEAR'][element] >= 1990:
            data_copy_80s = data_copy_80s.drop([element], axis=0)
    budget80 = data_copy_80s['BUDGET']
    gross80 = data_copy_80s['GROSS']
    plt.scatter(budget80, gross80, s=20, color='#ff9999')
    chart_sheet['U165'] = len(data_copy_80s)

    chart_sheet['T166'] = '90s'
    data_copy_90s = cinema_csv_file
    for element in range(0, len(data)):
        if element in data_copy_90s['YEAR'] and \
                (data_copy_90s['YEAR'][element] >= 2000 or data_copy_90s['YEAR'][element] < 1990):
            data_copy_90s = data_copy_90s.drop([element], axis=0)
    budget90 = data_copy_90s['BUDGET']
    gross90 = data_copy_90s['GROSS']
    plt.scatter(budget90, gross90, s=20, color='#66b3ff')
    chart_sheet['U166'] = len(data_copy_90s)

    chart_sheet['T167'] = '00s'
    data_copy_00s = cinema_csv_file
    for element in range(0, len(data)):
        if element in data_copy_00s['YEAR'] and \
                (data_copy_00s['YEAR'][element] >= 2010 or data_copy_00s['YEAR'][element] < 2000):
            data_copy_00s = data_copy_00s.drop([element], axis=0)
    budget00 = data_copy_00s['BUDGET']
    gross00 = data_copy_00s['GROSS']
    plt.scatter(budget00, gross00, s=20, color='#FFE873')
    chart_sheet['U167'] = len(data_copy_00s)

    chart_sheet['T168'] = '10s'
    data_copy_10s = cinema_csv_file
    for element in range(0, len(data)):
        if element in data_copy_10s['YEAR'] and \
                (data_copy_10s['YEAR'][element] >= 2020 or data_copy_10s['YEAR'][element] < 2010):
            data_copy_10s = data_copy_10s.drop([element], axis=0)
    budget10 = data_copy_10s['BUDGET']
    gross10 = data_copy_10s['GROSS']
    plt.scatter(budget10, gross10, s=20, color='#99ff99')
    chart_sheet['U168'] = len(data_copy_10s)

    chart_sheet['T169'] = '20s'
    data_copy_20s = cinema_csv_file
    for element in range(0, len(data)):
        if element in data_copy_20s['YEAR'] and data_copy_20s['YEAR'][element] < 2020:
            data_copy_20s = data_copy_20s.drop([element], axis=0)
    budget20 = data_copy_20s['BUDGET']
    gross20 = data_copy_20s['GROSS']
    plt.scatter(budget20, gross20, s=20, color='#ffcc99')
    chart_sheet['U169'] = len(data_copy_20s)

    chart_sheet['T170'] = 'TOTAL'
    chart_sheet['U170'] = len(data_copy_80s) + len(data_copy_90s) + len(data_copy_00s) + len(data_copy_10s) + \
                          len(data_copy_20s)

    plt.grid()
    plt.legend(('80s', '90s', '00s', '10s', '20s'))
    plt.xlabel("Budget")
    plt.ylabel("Gross")
    save_and_insert_image(chart_sheet, 'Budget-to-Gross Ratio by decade', 'J160')

    # 11) Budget-to-Gross Ratio
    plt.figure()
    plt.scatter(budget, gross, s=20)
    plt.grid()
    plt.xlabel("Budget")
    plt.ylabel("Gross")
    plt.axvline(stats.mean(budget), color="red", linewidth=1, linestyle="dashed")
    plt.axvline(stats.median(budget), color="red", linewidth=1, linestyle="dashed")
    plt.axhline(stats.mean(gross), color="green", linewidth=1, linestyle="dashed")
    plt.axhline(stats.median(gross), color="green", linewidth=1, linestyle="dashed")
    save_and_insert_image(chart_sheet, 'Budget-to-Gross Ratio', 'A186')

    chart_sheet['K193'] = 'BUDGET'
    chart_sheet['L193'] = 'GROSS'
    chart_sheet['J194'] = 'MEAN'
    chart_sheet['J195'] = 'MEDIAN'
    chart_sheet['K194'] = stats.mean(budget)
    chart_sheet['K195'] = stats.median(budget)
    chart_sheet['L194'] = stats.mean(gross)
    chart_sheet['L195'] = stats.median(gross)

    # 12) Linear Regression Simple All movies
    formula = 'GROSS~BUDGET'
    plt.figure()
    lm = smf.ols(formula=formula, data=cinema_csv_file).fit()
    expected_gross = lm.predict(pd.DataFrame(cinema_csv_file["BUDGET"]))
    plt.scatter(budget, gross, s=20)
    plt.grid()
    plt.xlabel("Budget")
    plt.ylabel("Gross")
    plt.plot(pd.DataFrame(cinema_csv_file["BUDGET"]), expected_gross, c="red", linewidth=2)
    save_and_insert_image(chart_sheet, 'Linear Regression Simple All movies', 'A212')

    # 13) Separated Budget-to-Gross Ratio by decade
    plt.figure()
    fig, (ax1, ax2, ax3, ax4, ax5) = plt.subplots(nrows=1, ncols=5, sharex='all', sharey='all', figsize=(20, 4.8))
    ax1.scatter(budget80, gross80, color='#ff9999')
    ax1.set_title('80s')
    ax1.grid()
    ax2.scatter(budget90, gross90, color='#66b3ff')
    ax2.set_title('90s')
    ax2.grid()
    ax3.scatter(budget00, gross00, color='#FFE873')
    ax3.set_title('00s')
    ax3.grid()
    ax4.scatter(budget10, gross10, color='#99ff99')
    ax4.set_title('10s')
    ax4.grid()
    ax5.scatter(budget20, gross20, color='#ffcc99')
    ax5.set_title('20s')
    ax5.grid()
    save_and_insert_image(chart_sheet, 'Separated Budget-to-Gross Ratio by decade', 'A238')

    # 14) Linear Regression Simple separated decades
    plt.figure()
    fig, (ax1, ax2, ax3, ax4, ax5) = plt.subplots(nrows=1, ncols=5, sharex='all', sharey='all', figsize=(20, 4.8))
    lm80 = smf.ols(formula=formula, data=data_copy_80s).fit()
    expected_gross80 = lm80.predict(pd.DataFrame(data_copy_80s["BUDGET"]))
    ax1.scatter(budget80, gross80, color='#ff9999')
    ax1.set_title('80s')
    ax1.grid()
    ax1.plot(pd.DataFrame(data_copy_80s["BUDGET"]), expected_gross80, c="red", linewidth=2)

    lm90 = smf.ols(formula=formula, data=data_copy_90s).fit()
    expected_gross90 = lm90.predict(pd.DataFrame(data_copy_90s["BUDGET"]))
    ax2.scatter(budget90, gross90, color='#66b3ff')
    ax2.set_title('90s')
    ax2.grid()
    ax2.plot(pd.DataFrame(data_copy_90s["BUDGET"]), expected_gross90, c="red", linewidth=2)

    lm00 = smf.ols(formula=formula, data=data_copy_00s).fit()
    expected_gross00 = lm00.predict(pd.DataFrame(data_copy_00s["BUDGET"]))
    ax3.scatter(budget00, gross00, color='#FFE873')
    ax3.set_title('00s')
    ax3.grid()
    ax3.plot(pd.DataFrame(data_copy_00s["BUDGET"]), expected_gross00, c="red", linewidth=2)

    lm10 = smf.ols(formula=formula, data=data_copy_10s).fit()
    expected_gross10 = lm10.predict(pd.DataFrame(data_copy_10s["BUDGET"]))
    ax4.scatter(budget10, gross10, color='#99ff99')
    ax4.set_title('10s')
    ax4.grid()
    ax4.plot(pd.DataFrame(data_copy_10s["BUDGET"]), expected_gross10, c="red", linewidth=2)

    lm20 = smf.ols(formula=formula, data=data_copy_20s).fit()
    expected_gross20 = lm20.predict(pd.DataFrame(data_copy_20s["BUDGET"]))
    ax5.scatter(budget20, gross20, color='#ffcc99')
    ax5.set_title('20s')
    ax5.grid()
    ax5.plot(pd.DataFrame(data_copy_20s["BUDGET"]), expected_gross20, c="red", linewidth=2)
    save_and_insert_image(chart_sheet, 'Linear Regression Simple separated decades', 'A264')

    # 15) 10 highest grossing films (treemap)
    plt.figure()
    chart_sheet['B290'] = 'Top 10 highest grossing films'
    data_highest_grossing = data.sort_values(['GROSS'], ascending=False)
    data_highest_grossing = data_highest_grossing.iloc[0:10]
    highest_grossing_title = data_highest_grossing['TITLE']
    highest_grossing_gross = data_highest_grossing['GROSS']
    norm = matplotlib.colors.Normalize(vmin=min(highest_grossing_gross), vmax=max(highest_grossing_gross))
    colors_grossing = [matplotlib.cm.Blues(norm(value)) for value in highest_grossing_gross]
    plt.axis('off')
    plt.rc('font', size=5)
    squarify.plot(sizes=highest_grossing_gross, label=highest_grossing_title, color=colors_grossing, alpha=0.8)
    save_and_insert_image(chart_sheet, 'Top 10 highest grossing films', 'A291')

    # 16) 10 highest rated movies (treemap)
    plt.figure()
    chart_sheet['M290'] = 'Top 10 highest rated movies'
    data_highest_rated = data.sort_values(['SCORE'], ascending=False)
    data_highest_rated = data_highest_rated.iloc[0:10]
    highest_rated_title = data_highest_rated['TITLE']
    highest_rated_score = data_highest_rated['SCORE']
    norm = matplotlib.colors.Normalize(vmin=min(highest_rated_score), vmax=max(highest_rated_score))
    colors_rated = [matplotlib.cm.Blues(norm(value)) for value in highest_rated_score]
    plt.axis('off')
    plt.rc('font', size=5)
    squarify.plot(sizes=highest_rated_score, label=highest_rated_title, color=colors_rated, alpha=0.8)
    save_and_insert_image(chart_sheet, 'Top 10 highest rated movies', 'J291')

    money_cells = ['K33', 'K34', 'L33', 'L34', 'K194', 'K195', 'L194', 'L195']
    for pos in money_cells:
        chart_sheet[pos].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2

    border_cell(chart_sheet, 'J3:K5')
    border_cell(chart_sheet, 'J31:L35')
    border_cell(chart_sheet, 'S56:U71')
    border_cell(chart_sheet, 'S107:AA127')
    border_cell(chart_sheet, 'K141:L152')
    border_cell(chart_sheet, 'T164:U170')
    border_cell(chart_sheet, 'J193:L195')

    merge_cells(chart_sheet, 'S57:S59')
    merge_cells(chart_sheet, 'S60:S62')
    merge_cells(chart_sheet, 'S63:S65')
    merge_cells(chart_sheet, 'S66:S68')
    merge_cells(chart_sheet, 'S69:S71')

    workbook.save(excel_file)

    for element in listdir('.'):
        if '.png' in element:
            remove(element)
    end = datetime.now()
    time = end - start
    print('The process took ' + str(time))
